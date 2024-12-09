import os
import json
import base64
import tempfile
import requests # not in aws runtime
import time
from io import BytesIO
from openpyxl import load_workbook # not in aws runtime

def authenticate_request(body, headers):
  ALLOW_IP_BASED_ACCESS = False
  contentLength = int(headers.get('Content-Length'))
  authorization = headers.get('authorization', None)
  requestIp = headers.get('X-Forwarded-For')
  apiKeyEnvStr = os.getenv('API_KEY')
  ipWhitelistEnvStr = os.getenv('IP_WHITELIST')
  ipWhitelist = ipWhitelistEnvStr.split(",")
  # block access if payload is empty
  if body == None or contentLength == 0:
    return {
      "statusCode": 422,
      "body": json.dumps({"message": "please provide a payload. Received empty request body."})
    }
  # block access if authorization header does not include the API TOKEN
  if authorization == None or apiKeyEnvStr == None or authorization != apiKeyEnvStr:
    return {
      "statusCode": 403,
      "body": json.dumps({"message": "Invalid authorization header."})
    }
  # block access if requesting IP-address is not whitelisted
  if ipWhitelistEnvStr == None or ipWhitelistEnvStr == "":
    return {
      "statusCode": 422,
      "body": json.dumps({"message": "ip address whitelist not found among environment variables."})
    }
  if len(ipWhitelist) == 1 and ipWhitelist[0] == "0.0.0.0":
    ALLOW_IP_BASED_ACCESS = True
  elif len(ipWhitelist) > 1:
    for ip in ipWhitelist:
      if requestIp == ip:
        ALLOW_IP_BASED_ACCESS = True
  if not ALLOW_IP_BASED_ACCESS:
    return {
      "statusCode": 403,
      "body": json.dumps({"message": "IP based access denied."})
    }
  # Everything is fine. Proceed
  return {
      "statusCode": 200
    }

def get_debug_info(event, decodedBody, headers):
  extracedEventKeys = {
    "path" : event["path"],
    "forwardedFor" : headers.get('X-Forwarded-For'),
    "decodedBody" : decodedBody
  }
  return {"statusCode": 200,
          "body": json.dumps({
            "eventKeys": extracedEventKeys
          })}

def logTimePassed(start_time, log_msg):
  print(f"{round((time.time_ns() - start_time) / 1000000)}ms time passed after [{log_msg}]")

def handler(event, context):
  start_time = time.time_ns()
  body = event["body"]
  headers = event["headers"]
  auth_response = authenticate_request(body, headers)
  if auth_response["statusCode"] != 200:
    return auth_response
  logTimePassed(start_time, "authenticate_request")
  decodedBytes = base64.b64decode(body)
  decodedBody = decodedBytes.decode("utf-8")
  debug_info = get_debug_info(event, decodedBody, headers)
  # return debug_info
  try:
    # Evaluate request body
    sheet_url = decodedBody
    if not sheet_url or "docs.google.com/spreadsheets" not in sheet_url:
      return {
        "statusCode": 400,
        "body": json.dumps({"error": "Missing spreadsheets url in body."})
      }
    if "pub?output=xlsx" in sheet_url or "export?format=xlsx" in sheet_url:
      pass
    elif "/edit" in sheet_url:
      sheet_url = sheet_url.split("/edit")[0] + "/export?format=xlsx"
    elif "/view" in sheet_url:
      sheet_url = sheet_url.split("/view")[0] + "/export?format=xlsx"
    elif "/pubhtml" in sheet_url:
      sheet_url = sheet_url.split("/pubhtml")[0] + "/pub?output=xlsx"

    # Download the file
    response = requests.get(sheet_url)
    logTimePassed(start_time, "sheet request")
    if response.status_code != 200:
      return {
        "statusCode": 500,
        "body": json.dumps({"error": "Failed to download the sheet"})
      }
    # Save to a temporary file on disk for potential S3 backup
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(response.content)
    logTimePassed(start_time, "temp file write")
    # Load Workbook directly from memory for performance reasons
    file_content = BytesIO(response.content)
    workbook = load_workbook(file_content, read_only=True)
    logTimePassed(start_time, "load workbook into var")
    sheet_names = workbook.sheetnames
    logTimePassed(start_time, "sheet name extraction")
    return {
      "statusCode": 200,
      "body": json.dumps({"sheets": sheet_names})
    }

  except Exception as e:
    return {
        "statusCode": 500,
        "body": json.dumps({"error": str(e)})
    }
